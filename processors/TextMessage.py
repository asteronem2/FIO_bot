import os
import re
from typing import NoReturn

from aiogram.types import InlineKeyboardButton as IButton, FSInputFile
from openpyxl import Workbook

import utils
from Bot import MsgModel
from Interfaces import TextMessageInterface


class StartMsg(TextMessageInterface):
    async def define(self) -> NoReturn | dict:
        rres = re.fullmatch(r'/start( [^ ]+|)', self.text_low)
        if rres:
            ref = None
            if rres.group(1):
                ref = rres.group(1).strip()
            return {'ref': ref}

    async def process(self, *args, **kwargs) -> None:
        await self.bot.send_message(MsgModel(
            chat_id=self.chat.id,
            text=self.locale['StartTxt'],
            markup=[[IButton(text='Принять участие', callback_data='take_part')]]
        ))
        ref = None
        if kwargs.get('ref'):
            ref = kwargs['ref']

        await self.user_core.update({'id': self.db_user.id}, next_message_info=None, ref=ref)


class GetTable(TextMessageInterface):
    async def define(self) -> NoReturn | dict:
        if self.user.id == utils.ADMIN_ID:
            rres = re.fullmatch(r'таблица', self.text_low)
            if rres:
                return True

    async def process(self, *args, **kwargs) -> None:
        file_name = 'users_table.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Выгрузка базы данных'

        res = await self.user_core.find_all()

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30

        ws.cell(row=1, column=1, value='ID')
        ws.cell(row=1, column=2, value='USER_ID')
        ws.cell(row=1, column=3, value='USERNAME')
        ws.cell(row=1, column=4, value='FIRST_NAME')
        ws.cell(row=1, column=5, value='ФИО')

        row = 2

        for i in res:
            ws.cell(row=row, column=1, value=i.id)
            ws.cell(row=row, column=2, value=i.user_id)
            ws.cell(row=row, column=3, value=i.username or ' ')
            ws.cell(row=row, column=4, value=i.first_name or ' ')
            ws.cell(row=row, column=5, value=i.fio or ' ')
            row += 1

        wb.save(file_name)

        await self.bot.bot.send_document(self.chat.id, FSInputFile(file_name))
        os.remove(file_name)


class Distribution(TextMessageInterface):
    async def define(self) -> NoReturn | dict:
        if self.user.id == utils.ADMIN_ID:
            rres = re.fullmatch(r'рассылка', self.text_low)
            if rres:
                return True

    async def process(self, *args, **kwargs) -> None:
        await self.bot.send_message(MsgModel(
            chat_id=self.user.id,
            text=self.locale['DistributionTxt']
        ))
        await self.user_core.update({'id': self.db_user.id}, next_message_info='distribution')


class Referrals(TextMessageInterface):
    async def define(self) -> NoReturn | dict:
        rres = re.fullmatch(r'рефералы', self.text_low)
        if rres:
            return True

    async def process(self, *args, **kwargs) -> None:
        all_user = await self.user_core.find_all()

        all_ref_types = {'None': 0}

        for i in all_user:
            if i.ref:
                if i.ref in list(all_ref_types.keys()):
                    all_ref_types[i.ref] += 1
                else:
                    all_ref_types[i.ref] = 1
            else:
                all_ref_types['None'] += 1

        final_text = f'<b>Всего: {len(all_user)}</b>\n\n'
        final_text += 'Без тега: ' + str(all_ref_types['None'])

        all_ref_types.pop('None')

        for ref, count in all_ref_types.items():
            final_text += '\n'
            final_text += f'{ref}: {count}'

        await self.bot.send_message(MsgModel(
            chat_id=self.chat.id,
            text=final_text
        ))


class NewRefLink(TextMessageInterface):
    async def define(self) -> NoReturn | dict:
        rres = re.fullmatch(r'ссылка ([a-z0-9]+)', self.text_low)
        if rres:
            ref = rres.group(1)
            return {'ref': ref}

    async def process(self, *args, **kwargs) -> None:
        me_username = (await self.bot.bot.get_me()).username
        ref = kwargs['ref']

        await self.bot.send_message(MsgModel(
            chat_id=self.chat.id,
            text=f'Новая реферальная ссылка:\n<code>{"https://t.me/" + me_username + "?start=" + ref}</code>'
        ))
